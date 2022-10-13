import openpyxl
import json


class food_data():
    excel_file = 'staticfiles/foodDB99.xlsx'
    json_file = 'staticfiles/food_data.json'
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    
    sheet = wb.worksheets[0]
    
    key_list = []
    for col_num in range(1, sheet.max_column + 1):
        key_list.append(sheet.cell(row=1, column=col_num).value)
    
    food_dict = {}
    key_index = 3
    for row_num in range(2, sheet.max_row + 1):
        tmp_dict = {}
        for col_num in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row_num, column=col_num).value
            tmp_dict[key_list[col_num - 1]] = val
        
        food_dict[tmp_dict[key_list[key_index]]] = tmp_dict
    wb.close()

    with open(json_file, 'w', encoding='utf-8') as fp:
        json.dump(food_dict, fp, indent=4, ensure_ascii=False)


        
# class fisheries_data():
#     excel_file = 'staticfiles/통합_식품영양성분DB_수산물_20221011.xlsx'
#     json_file = 'staticfiles/fisheries_data.json'
#     wb = openpyxl.load_workbook(excel_file, read_only=True)
    
#     sheet = wb.worksheets[0]
    
#     key_list = []
#     for col_num in range(1, sheet.max_column + 1):
#         key_list.append(sheet.cell(row=1, column=col_num).value)
    
#     fish_dict = {}
#     key_index = 5
#     for row_num in range(2, sheet.max_row + 1):
#         tmp_dict = {}
#         for col_num in range(1, sheet.max_column + 1):
#             val = sheet.cell(row=row_num, column=col_num).value
#             tmp_dict[key_list[col_num - 1]] = val
        
#         fish_dict[tmp_dict[key_list[key_index]]] = tmp_dict
    
#     wb.close()

#     with open(json_file, 'w', encoding='utf-8') as fp:
#         json.dump(fish_dict, fp, indent=4, ensure_ascii=False)
        
        
        
        
        
# class agriculture_data():
#     excel_file = 'staticfiles/통합_식품영양성분DB_농축산물_20221011.xlsx'
#     json_file = 'staticfiles/agriculture_data.json'
#     wb = openpyxl.load_workbook(excel_file, read_only=True)
    
#     sheet = wb.worksheets[0]
    
#     key_list = []
#     for col_num in range(1, sheet.max_column + 1):
#         key_list.append(sheet.cell(row=1, column=col_num).value)
    
#     agriculture_dict = {}
#     key_index = 5
#     for row_num in range(2, sheet.max_row + 1):
#         tmp_dict = {}
#         for col_num in range(1, sheet.max_column + 1):
#             val = sheet.cell(row=row_num, column=col_num).value
#             tmp_dict[key_list[col_num - 1]] = val
        
#         agriculture_dict[tmp_dict[key_list[key_index]]] = tmp_dict
    
#     wb.close()

#     with open(json_file, 'w', encoding='utf-8') as fp:
#         json.dump(agriculture_dict, fp, indent=4, ensure_ascii=False)
              
        
        
# class processed_data():
#     excel_file = 'staticfiles/통합_식품영양성분DB_가공식품_20221011.xlsx'
#     json_file = 'staticfiles/processed.json'
#     wb = openpyxl.load_workbook(excel_file, read_only=True)
    
#     sheet = wb.worksheets[0]
    
#     key_list = []
#     for col_num in range(1, sheet.max_column + 1):
#         key_list.append(sheet.cell(row=1, column=col_num).value)
    
#     processed_dict = {}
#     key_index = 5
#     for row_num in range(2, sheet.max_row + 1):
#         tmp_dict = {}
#         for col_num in range(1, sheet.max_column + 1):
#             val = sheet.cell(row=row_num, column=col_num).value
#             tmp_dict[key_list[col_num - 1]] = val
            
#         processed_dict[tmp_dict[key_list[key_index]]] = tmp_dict
    
#     wb.close()

#     with open(json_file, 'w', encoding='utf-8') as fp:
#         json.dump(processed_dict, fp, indent=4, ensure_ascii=False)






    